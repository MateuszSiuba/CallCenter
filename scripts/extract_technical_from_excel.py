import sys
import json
from openpyxl import load_workbook

if len(sys.argv) < 3:
    print('Usage: extract_technical_from_excel.py <excel_path> <missing_json_path> [out_json_path]')
    sys.exit(2)

excel_path = sys.argv[1]
missing_path = sys.argv[2]
out_path = None
if len(sys.argv) >= 4:
    out_path = sys.argv[3]

with open(missing_path, 'r', encoding='utf8') as f:
    missing = json.load(f)
missing_set = set(missing)

wb = load_workbook(filename=excel_path, data_only=True)
results = {}

# helper: normalize cell value
def norm(v):
    if v is None:
        return ''
    return str(v).strip()

for sheet in wb.worksheets:
    # try to find header row: first row with more than 1 non-empty cell
    max_row = sheet.max_row
    max_col = sheet.max_column
    header_row_idx = None
    for r in range(1, min(6, max_row)+1):
        values = [norm(sheet.cell(row=r, column=c).value) for c in range(1, max_col+1)]
        non_empty = sum(1 for v in values if v)
        if non_empty >= 2:
            header_row_idx = r
            headers = values
            break
    if header_row_idx is None:
        continue

    # build rows
    for r in range(header_row_idx+1, max_row+1):
        row_vals = [norm(sheet.cell(row=r, column=c).value) for c in range(1, max_col+1)]
        if all(v=='' for v in row_vals):
            continue
        # attempt to detect model column: any cell in row that equals a missing model or contains it
        model_cell_index = None
        model_name = None
        for i,v in enumerate(row_vals):
            for m in missing_set:
                if not m: continue
                if v == m or m in v:
                    model_cell_index = i
                    model_name = m
                    break
            if model_cell_index is not None:
                break
        if model_cell_index is None:
            # sometimes model is in combined column or separate label; skip
            continue
        # build technicalFlat mapping from other headers
        tech = {}
        for i,h in enumerate(headers):
            key = norm(h)
            if not key: continue
            if i == model_cell_index: continue
            val = row_vals[i]
            if val:
                tech[f"{key}"] = val
        if tech:
            # merge if duplicate
            existing = results.get(model_name, {})
            existing.update(tech)
            results[model_name] = existing

# write output
if out_path:
    with open(out_path, 'w', encoding='utf8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
else:
    print(json.dumps(results, ensure_ascii=False, indent=2))
